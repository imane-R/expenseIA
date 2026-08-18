"""Chargement fidèle du fichier Excel brut dans la table de staging."""

from __future__ import annotations

import logging
import math
import warnings
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.exc import SQLAlchemyError

from database.connection import create_db_engine


logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL_PATH = (
    PROJECT_ROOT / "data" / "raw" / "expenses data-20260722102438.xlsx"
)
SOURCE_SHEET = "data"
EXPECTED_SOURCE_ROWS = 7_071

SOURCE_TO_STAGING = {
    "Numéro (Dépense)": "expense_number",
    "Date": "expense_date",
    "Nom (Dépense)": "expense_name",
    "Type": "expense_type",
    "Montant TTC devise système": "amount_ttc_system_currency",
    "Facturable": "billable",
    "Code projet": "project_code",
    "Nom (Projet)": "project_name",
    "Taux de taxe": "tax_rate",
    "Montant HT devise système": "amount_ht_system_currency",
    "Statut": "status",
    "Nom de fichier (Justificatif)": "receipt_filename",
    "Date d'approbation": "approval_date",
    "Motif du refus": "rejection_reason",
}

INSERT_STAGING_SQL = text(
    """
    INSERT INTO staging_expenses_raw (
        source_file,
        source_row_number,
        expense_number,
        expense_date,
        expense_name,
        expense_type,
        amount_ttc_system_currency,
        billable,
        project_code,
        project_name,
        tax_rate,
        amount_ht_system_currency,
        status,
        receipt_filename,
        approval_date,
        rejection_reason
    ) VALUES (
        :source_file,
        :source_row_number,
        :expense_number,
        :expense_date,
        :expense_name,
        :expense_type,
        :amount_ttc_system_currency,
        :billable,
        :project_code,
        :project_name,
        :tax_rate,
        :amount_ht_system_currency,
        :status,
        :receipt_filename,
        :approval_date,
        :rejection_reason
    )
    ON CONFLICT (source_file, source_row_number) DO NOTHING
    """
)

COUNT_STAGING_SQL = text(
    """
    SELECT COUNT(*)
    FROM staging_expenses_raw
    WHERE source_file = :source_file
    """
)


@dataclass(frozen=True)
class StagingImportSummary:
    """Résultats du chargement brut pour un fichier source."""

    source_rows: int
    rows_before: int
    inserted_rows: int
    rows_after: int


def _raw_value_to_text(value: object) -> str | None:
    """Sérialise une cellule sans appliquer de règle métier ou ML."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if math.isnan(value):
            return None
        return str(value)
    return str(value)


def read_source_rows(excel_path: Path = DEFAULT_EXCEL_PATH) -> list[dict[str, object]]:
    """Lit les 14 colonnes métier sans modifier le classeur Excel."""
    if not excel_path.exists():
        raise FileNotFoundError(f"Fichier Excel introuvable : {excel_path}")

    warnings.filterwarnings(
        "ignore", message="Workbook contains no default style", module="openpyxl"
    )
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        if SOURCE_SHEET not in workbook.sheetnames:
            raise ValueError(f"Feuille Excel absente : {SOURCE_SHEET}")
        worksheet = workbook[SOURCE_SHEET]

        header_row = list(
            next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        )
        expected_headers = list(SOURCE_TO_STAGING)
        if header_row[: len(expected_headers)] != expected_headers:
            raise ValueError(
                "Les 14 colonnes de la feuille data ne correspondent pas au schéma attendu."
            )
        if any(value is not None for value in header_row[len(expected_headers) :]):
            raise ValueError("Une colonne métier inattendue existe après les 14 colonnes.")

        records: list[dict[str, object]] = []
        for source_row_number, row in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            business_values = row[: len(expected_headers)]
            if any(value is not None for value in row[len(expected_headers) :]):
                raise ValueError(
                    f"Valeur inattendue après la 14e colonne à la ligne "
                    f"{source_row_number}."
                )

            record: dict[str, object] = {
                "source_file": excel_path.name,
                "source_row_number": source_row_number,
            }
            for source_column, value in zip(
                expected_headers, business_values, strict=True
            ):
                staging_column = SOURCE_TO_STAGING[source_column]
                record[staging_column] = _raw_value_to_text(value)
            records.append(record)
    finally:
        workbook.close()

    if len(records) != EXPECTED_SOURCE_ROWS:
        raise ValueError(
            f"Le fichier contient {len(records)} lignes de données au lieu de "
            f"{EXPECTED_SOURCE_ROWS}."
        )
    return records


def load_staging(excel_path: Path = DEFAULT_EXCEL_PATH) -> StagingImportSummary:
    """Insère le brut dans une transaction, sans dupliquer une ligne source."""
    records = read_source_rows(excel_path)
    source_file = excel_path.name
    engine = create_db_engine()

    try:
        with engine.begin() as connection:
            # Sérialise les imports de staging pour éviter deux chargements concurrents.
            connection.execute(
                text(
                    "SELECT pg_advisory_xact_lock("
                    "CAST(hashtext(:lock_name) AS BIGINT))"
                ),
                {"lock_name": f"expenseai_load_staging:{source_file}"},
            )
            rows_before = int(
                connection.execute(
                    COUNT_STAGING_SQL, {"source_file": source_file}
                ).scalar_one()
            )

            connection.execute(INSERT_STAGING_SQL, records)

            rows_after = int(
                connection.execute(
                    COUNT_STAGING_SQL, {"source_file": source_file}
                ).scalar_one()
            )
            inserted_rows = rows_after - rows_before

            if rows_after != EXPECTED_SOURCE_ROWS:
                raise RuntimeError(
                    f"La staging contient {rows_after} lignes pour {source_file}, "
                    f"au lieu de {EXPECTED_SOURCE_ROWS}."
                )
    except (SQLAlchemyError, RuntimeError):
        logger.exception(
            "Chargement staging annulé : la transaction PostgreSQL a été rollbackée."
        )
        raise

    return StagingImportSummary(
        source_rows=len(records),
        rows_before=rows_before,
        inserted_rows=inserted_rows,
        rows_after=rows_after,
    )


def print_summary(summary: StagingImportSummary) -> None:
    """Affiche les compteurs de contrôle demandés."""
    print(f"Lignes lues dans le fichier source : {summary.source_rows}")
    print(f"Lignes déjà présentes avant import : {summary.rows_before}")
    print(f"Lignes nouvellement insérées : {summary.inserted_rows}")
    print(f"Lignes présentes après import : {summary.rows_after}")


def main() -> int:
    """Point d'entrée en ligne de commande."""
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    try:
        summary = load_staging()
    except (FileNotFoundError, SQLAlchemyError, ValueError, RuntimeError) as exc:
        logger.error("Échec du chargement staging : %s", exc)
        return 1
    print_summary(summary)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
